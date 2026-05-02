import openpyxl

wb = openpyxl.load_workbook('Onboarding Info/Activation Questions UK-IRE.xlsx', data_only=True)
print(f'📊 Sheets: {", ".join(wb.sheetnames)}\n')

for sheet_name in wb.sheetnames:
    print('\n' + '='*60)
    print(f'SHEET: {sheet_name}')
    print('='*60)
    ws = wb[sheet_name]

    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
        non_empty = [str(cell).strip() for cell in row if cell and str(cell).strip() and str(cell).strip() != 'None']

        if not non_empty:
            continue

        first_col = str(row[0]).strip() if row[0] else ''
        rest_cols = [cell for cell in row[1:] if cell and str(cell).strip() and str(cell).strip() != 'None']

        if first_col and len(rest_cols) == 0:
            print(f'\n## {first_col}')
        elif non_empty:
            print(f'  • {" | ".join(non_empty)}')

print('\n' + '='*60)
print('SUMMARY')
print('='*60)
print(f'Total sheets: {len(wb.sheetnames)}')
